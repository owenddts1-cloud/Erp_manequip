import openpyxl
import uuid
import re

# Tech map from names to email
TECH_MAP = {
    'ALDEMAR': ('aldemar@manequip.com', 'Aldemar Técnico'),
    'DANIEL': ('daniel@manequip.com', 'Daniel Técnico'),
    'GENILSON': ('genilson@manequip.com', 'Genilson Técnico'),
    'HUGO': ('hugo@manequip.com', 'Hugo Técnico'),
    'LUAN': ('luan@manequip.com', 'Luan Técnico'),
    'MAIKE': ('maike@manequip.com', 'Maike Técnico'),
    'SAMIR': ('samir@manequip.com', 'Samir Técnico'),
    'VINICIUS': ('vinicius@manequip.com', 'Vinícius Técnico'),
    'VINÍCIUS': ('vinicius@manequip.com', 'Vinícius Técnico'),
    'WENDEL': ('wendel@manequip.com', 'Wendel Técnico'),
    'GUILHERME / LUAN': ('luan@manequip.com', 'Luan Técnico'), # default to Luan
    'GUILHERME': ('guilherme@manequip.com', 'Guilherme Técnico'),
}

def clean_patrimonio(p):
    if p is None:
        return 'NULL'
    p_str = str(p).strip()
    if p_str in ('', 'None', '—', '-', 'NT'):
        return 'NULL'
    # Clean non-alphanumeric chars if necessary, but tag_id is text so it's fine
    return f"'{p_str}'"

def clean_str(s):
    if s is None:
        return 'NULL'
    s_str = str(s).strip().replace("'", "''")
    if s_str in ('', 'None', '—', '-'):
        return 'NULL'
    return f"'{s_str}'"

def normalize_periodicity(p):
    if not p:
        return 'Mensal'
    p = str(p).strip().upper()
    if 'SEMANAL' in p:
        return 'Semanal'
    elif 'MENSAL' in p:
        return 'Mensal'
    elif 'BIMESTRAL' in p:
        return 'Bimestral'
    elif 'TRIMESTRAL' in p or 'TRIMESTARL' in p:
        return 'Trimestral'
    elif 'SEMESTRAL' in p:
        return 'Semestral'
    elif 'ANUAL' in p:
        return 'Anual'
    return 'Mensal'

def generate_sql():
    sql_lines = []
    
    # 1. Header and transactions
    sql_lines.append("-- ==========================================================")
    sql_lines.append("-- AUTOMATICALLY GENERATED PREVENTIVE MAINTENANCE DATA IMPORT")
    sql_lines.append("-- ==========================================================")
    sql_lines.append("BEGIN;")
    
    # 2. Helper functions to avoid duplication
    sql_lines.append("""
CREATE OR REPLACE FUNCTION public.get_or_create_ativo(
    p_tag_id text,
    p_nome text,
    p_setor text,
    p_criticidade text,
    p_status text
) RETURNS uuid AS $$
DECLARE
    v_id uuid;
BEGIN
    -- Try to find by tag_id if not null
    IF p_tag_id IS NOT NULL AND p_tag_id <> '' THEN
        SELECT id INTO v_id FROM public.ativos WHERE tag_id = p_tag_id LIMIT 1;
    ELSE
        -- If tag_id is null, find by name and sector
        SELECT id INTO v_id FROM public.ativos WHERE nome = p_nome AND (setor = p_setor OR (setor IS NULL AND p_setor IS NULL)) LIMIT 1;
    END IF;

    -- If not found, insert
    IF v_id IS NULL THEN
        INSERT INTO public.ativos (tag_id, nome, setor, criticidade, status)
        VALUES (p_tag_id, p_nome, p_setor, p_criticidade, p_status)
        RETURNING id INTO v_id;
    END IF;

    RETURN v_id;
END;
$$ LANGUAGE plpgsql SECURITY DEFINER;

CREATE OR REPLACE FUNCTION public.get_or_create_planejamento(
    p_ativo_id uuid,
    p_titulo text,
    p_descricao text,
    p_periodicidade text,
    p_meses_execucao integer[]
) RETURNS uuid AS $$
DECLARE
    v_id uuid;
BEGIN
    SELECT id INTO v_id 
    FROM public.preventivas_planejamento 
    WHERE ativo_id = p_ativo_id AND titulo = p_titulo AND periodicidade = p_periodicidade 
    LIMIT 1;

    IF v_id IS NULL THEN
        INSERT INTO public.preventivas_planejamento (ativo_id, titulo, descricao, periodicidade, meses_execucao)
        VALUES (p_ativo_id, p_titulo, p_descricao, p_periodicidade, p_meses_execucao)
        RETURNING id INTO v_id;
    ELSE
        -- Update months if it exists
        UPDATE public.preventivas_planejamento
        SET meses_execucao = p_meses_execucao,
            descricao = COALESCE(p_descricao, descricao)
        WHERE id = v_id;
    END IF;

    RETURN v_id;
END;
$$ LANGUAGE plpgsql SECURITY DEFINER;

CREATE OR REPLACE FUNCTION public.upsert_preventiva_mensal(
    p_planejamento_id uuid,
    p_ativo_id uuid,
    p_titulo text,
    p_descricao text,
    p_mes integer,
    p_ano integer,
    p_tecnico_email text,
    p_status text,
    p_dia_execucao integer
) RETURNS uuid AS $$
DECLARE
    v_tecnico_id uuid;
    v_id uuid;
    v_concluido_em timestamp with time zone := null;
    v_data_limite date;
BEGIN
    -- Resolve technician id
    IF p_tecnico_email IS NOT NULL THEN
        SELECT id INTO v_tecnico_id FROM public.profiles WHERE LOWER(email) = LOWER(p_tecnico_email) LIMIT 1;
    END IF;

    -- Calculate data_limite
    IF p_dia_execucao IS NOT NULL THEN
        v_data_limite := make_date(p_ano, p_mes, p_dia_execucao);
    ELSE
        v_data_limite := (date_trunc('month', make_date(p_ano, p_mes, 1)) + interval '1 month' - interval '1 day')::date;
    END IF;

    -- Set concluido_em if completed
    IF p_status = 'Concluído' THEN
        IF p_dia_execucao IS NOT NULL THEN
            v_concluido_em := make_timestamptz(p_ano, p_mes, p_dia_execucao, 12, 0, 0);
        ELSE
            v_concluido_em := NOW();
        END IF;
    END IF;

    -- Check if already exists for this month/year/asset
    SELECT id INTO v_id 
    FROM public.preventivas_mensais 
    WHERE ativo_id = p_ativo_id AND mes = p_mes AND ano = p_ano AND titulo = p_titulo
    LIMIT 1;

    IF v_id IS NULL THEN
        INSERT INTO public.preventivas_mensais (planejamento_id, ativo_id, titulo, descricao, mes, ano, tecnico_responsavel, status, data_limite, concluido_em)
        VALUES (p_planejamento_id, p_ativo_id, p_titulo, p_descricao, p_mes, p_ano, v_tecnico_id, p_status, v_data_limite, v_concluido_em)
        RETURNING id INTO v_id;
    ELSE
        UPDATE public.preventivas_mensais
        SET status = p_status,
            tecnico_responsavel = v_tecnico_id,
            concluido_em = v_concluido_em,
            data_limite = v_data_limite,
            descricao = COALESCE(p_descricao, descricao)
        WHERE id = v_id;
    END IF;

    RETURN v_id;
END;
$$ LANGUAGE plpgsql SECURITY DEFINER;
""")

    # 3. Create Technician Accounts
    sql_lines.append("\n-- 1. Seeding Technicians in Auth and Profiles")
    for clean_name, (email, name) in TECH_MAP.items():
        user_uuid = str(uuid.uuid4())
        sql_lines.append(f"""
-- Tech: {clean_name}
INSERT INTO auth.users (
    instance_id, id, aud, role, email, encrypted_password, email_confirmed_at, 
    recovery_sent_at, last_sign_in_at, raw_app_meta_data, raw_user_meta_data, 
    created_at, updated_at, confirmation_token, email_change, email_change_token_new, recovery_token
)
SELECT 
    '00000000-0000-0000-0000-000000000000',
    '{user_uuid}',
    'authenticated',
    'authenticated',
    '{email}',
    extensions.crypt('Manequip123!', extensions.gen_salt('bf')),
    current_timestamp,
    current_timestamp,
    current_timestamp,
    '{{"provider":"email","providers":["email"]}}',
    '{{"full_name":"{name}","job_title":"Técnico de Manutenção"}}',
    current_timestamp,
    current_timestamp,
    '', '', '', ''
WHERE NOT EXISTS (
    SELECT 1 FROM auth.users WHERE email = '{email}'
);

INSERT INTO public.profiles (id, email, full_name, role, is_approved)
SELECT id, email, '{name}', 'Técnico', true
FROM auth.users
WHERE email = '{email}'
ON CONFLICT (id) DO UPDATE
SET role = 'Técnico', is_approved = true, full_name = EXCLUDED.full_name;
""")

    # 4. Parse FO-SGI-032 and populate assets and plan templates
    print("Parsing planning sheet...")
    wb_plan = openpyxl.load_workbook(r"C:\Users\Manutenção\Downloads\FO-SGI-032 PLANEJAMENTO DE MANUTENCAO PREVENTIVA REV07 - 2026.xlsx", data_only=True)
    sheet_plan = wb_plan['FO-SGI-032']
    
    # Months columns index 7 to 18
    months_indices = list(range(7, 19))
    
    sql_lines.append("\n-- 2. Seeding Assets and Annual Planning Templates (218 rows)")
    
    plan_count = 0
    asset_id_var_map = {} # Map (name, patrimonio) to SQL variable name or UUID
    
    for r in range(6, sheet_plan.max_row + 1):
        row_vals = [cell.value for cell in sheet_plan[r]]
        if len(row_vals) < 6 or not row_vals[1]:
            continue
            
        equipamento = str(row_vals[1]).strip()
        patrimonio = row_vals[2]
        periodicidade = row_vals[3]
        localizacao = row_vals[4]
        operacao = row_vals[5]
        
        # Clean values
        pat_clean = clean_patrimonio(patrimonio)
        name_clean = clean_str(equipamento)
        setor_clean = clean_str(localizacao)
        per_clean = normalize_periodicity(periodicidade)
        
        # Determine planned months
        months_exec = []
        for i, col_idx in enumerate(months_indices):
            val = row_vals[col_idx]
            if val in ('P', 'R', 'p', 'r'):
                months_exec.append(i + 1)
        
        months_array_str = f"ARRAY{months_exec}" if months_exec else "'{}'::integer[]"
        
        # Define a SQL variable name for this planning/asset
        var_base = re.sub(r'[^a-zA-Z0-9]', '_', equipamento)[:30]
        var_asset = f"v_a_{var_base}_{r}"
        var_plan = f"v_p_{var_base}_{r}"
        
        # Generate SQL variables and function calls
        sql_lines.append(f"""
DO $$
DECLARE
    {var_asset} uuid;
    {var_plan} uuid;
BEGIN
    -- Get or create asset
    {var_asset} := public.get_or_create_ativo({pat_clean}, {name_clean}, {setor_clean}, 'Média', 'Operacional');
    
    -- Get or create planning template
    {var_plan} := public.get_or_create_planejamento(
        {var_asset},
        'Preventiva ' || {name_clean},
        'Realizar manutenção preventiva periódica conforme plano FO-SGI-032.',
        '{per_clean}',
        {months_array_str}
    );
END $$;""")
        plan_count += 1
        
    print(f"Generated SQL for {plan_count} planning templates.")

    # 5. Parse ABR2026 execution sheet
    print("Parsing execution sheet...")
    wb_exec = openpyxl.load_workbook(r"C:\Users\Manutenção\Downloads\PREVENTIVAS.xlsx", data_only=True)
    sheet_exec = wb_exec['ABR2026']
    
    sql_lines.append("\n-- 3. Seeding Monthly Execution Tasks for April 2026 (55 rows)")
    
    exec_count = 0
    for r in range(2, sheet_exec.max_row + 1):
        row_vals = [cell.value for cell in sheet_exec[r]]
        if len(row_vals) < 3 or not row_vals[2]:
            continue
            
        status = row_vals[0]
        chamado = row_vals[1]
        atividade = str(row_vals[2]).strip()
        patrimonio = row_vals[3]
        periodicidade = row_vals[4]
        localidade = row_vals[5]
        operacao = row_vals[6]
        
        # Clean values
        pat_clean = clean_patrimonio(patrimonio)
        name_clean = clean_str(atividade)
        setor_clean = clean_str(localidade)
        per_clean = normalize_periodicity(periodicidade)
        
        # Map status
        status_clean = 'Em aberto'
        if status == 'REALIZADA':
            status_clean = 'Concluído'
        elif status == 'EM ATENDIMENTO':
            status_clean = 'Em atendimento'
            
        # Find technician and day of execution
        tech_email = 'NULL'
        day_exec = 'NULL'
        
        for col_idx in range(7, 38):
            if col_idx < len(row_vals) and row_vals[col_idx] is not None:
                val = str(row_vals[col_idx]).strip()
                if val:
                    # Look up in TECH_MAP
                    tech_upper = val.upper()
                    if tech_upper in TECH_MAP:
                        tech_email = f"'{TECH_MAP[tech_upper][0]}'"
                    else:
                        # Fallback / cleaning
                        for key, (email, _) in TECH_MAP.items():
                            if key in tech_upper or tech_upper in key:
                                tech_email = f"'{email}'"
                                break
                    
                    # Day of month is the column header (from 1 to 31)
                    day_exec = int(col_idx - 6)
                    break
        
        # Write PL/pgSQL block to upsert
        var_base = re.sub(r'[^a-zA-Z0-9]', '_', atividade)[:30]
        var_asset = f"v_a_ex_{var_base}_{r}"
        var_plan = f"v_p_ex_{var_base}_{r}"
        
        sql_lines.append(f"""
DO $$
DECLARE
    {var_asset} uuid;
    {var_plan} uuid;
BEGIN
    -- Get or create asset
    {var_asset} := public.get_or_create_ativo({pat_clean}, {name_clean}, {setor_clean}, 'Média', 'Operacional');
    
    -- Try to find planning template by asset id and title
    SELECT id INTO {var_plan}
    FROM public.preventivas_planejamento
    WHERE ativo_id = {var_asset} AND titulo = 'Preventiva ' || {name_clean}
    LIMIT 1;
    
    -- If planning doesn't exist, create a default template
    IF {var_plan} IS NULL THEN
        {var_plan} := public.get_or_create_planejamento(
            {var_asset},
            'Preventiva ' || {name_clean},
            'Realizar manutenção preventiva periódica.',
            '{per_clean}',
            ARRAY[4] -- default execution in April
        );
    END IF;
    
    -- Upsert the monthly execution task
    PERFORM public.upsert_preventiva_mensal(
        {var_plan},
        {var_asset},
        'Preventiva ' || {name_clean},
        'Realizar manutenção preventiva periódica conforme plano de manutenção FO-SGI-032.',
        4, -- April
        2026, -- Year 2026
        {tech_email},
        '{status_clean}',
        {day_exec}
    );
END $$;""")
        exec_count += 1
        
    print(f"Generated SQL for {exec_count} execution tasks.")
    
    # 6. Commit transaction and cleanup
    sql_lines.append("\nCOMMIT;")
    sql_lines.append("\n-- Cleanup helper functions to keep schema clean")
    sql_lines.append("DROP FUNCTION IF EXISTS public.get_or_create_ativo(text, text, text, text, text);")
    sql_lines.append("DROP FUNCTION IF EXISTS public.get_or_create_planejamento(uuid, text, text, text, integer[]);")
    sql_lines.append("DROP FUNCTION IF EXISTS public.upsert_preventiva_mensal(uuid, uuid, text, text, integer, integer, text, text, integer);")
    
    # Save file
    sql_content = "\n".join(sql_lines)
    with open(r"c:\Users\Manutenção\MANEQUIP\scratch\import_data.sql", "w", encoding="utf-8") as f:
        f.write(sql_content)
        
    print("SQL generation complete! Saved to scratch/import_data.sql")

if __name__ == "__main__":
    generate_sql()
