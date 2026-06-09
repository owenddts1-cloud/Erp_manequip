import openpyxl
import json

def inspect_file(filepath):
    print(f"\n--- Inspecting {filepath} ---")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    print("Sheets:", wb.sheetnames)
    for name in wb.sheetnames[:2]:
        sheet = wb[name]
        print(f"\nSheet '{name}' dimensions: {sheet.dimensions}")
        # Print first 5 rows
        for i in range(1, min(10, sheet.max_row + 1)):
            row_values = [cell.value for cell in sheet[i]]
            print(f"Row {i}: {row_values[:20]}")

inspect_file(r"C:\Users\Manutenção\Downloads\FO-SGI-032 PLANEJAMENTO DE MANUTENCAO PREVENTIVA REV07 - 2026.xlsx")
inspect_file(r"C:\Users\Manutenção\Downloads\PREVENTIVAS.xlsx")
