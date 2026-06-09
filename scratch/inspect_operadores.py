import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\Manutenção\Downloads\FO-SGI-032 PLANEJAMENTO DE MANUTENCAO PREVENTIVA REV07 - 2026.xlsx", data_only=True)
if 'OPERADORES' in wb.sheetnames:
    sheet = wb['OPERADORES']
    print(f"\nSheet 'OPERADORES' dimensions: {sheet.dimensions}")
    for i in range(1, min(15, sheet.max_row + 1)):
        row_values = [cell.value for cell in sheet[i]]
        print(f"Row {i}: {row_values[:15]}")
else:
    print("OPERADORES sheet not found")
