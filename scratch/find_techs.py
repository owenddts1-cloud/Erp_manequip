import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\Manutenção\Downloads\PREVENTIVAS.xlsx", data_only=True)
for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    print(f"\n--- Searching in sheet '{sheetname}' ---")
    for r in range(1, min(20, sheet.max_row + 1)):
        row_vals = [cell.value for cell in sheet[r]]
        # Search for any known technician name
        for val in row_vals:
            if val and any(name in str(val).upper() for name in ["GENILSON", "ALDEMAR", "LUAN", "WENDEL", "HUGO", "VIN", "SAMIR"]):
                print(f"Row {r} contains tech name: {row_vals}")
                break
